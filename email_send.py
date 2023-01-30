import openpyxl
import smtplib
from email.mime.text import MIMEText


def sendemail(username, app_password, excel_sheet):
    # Load the Excel sheet
    wb = openpyxl.load_workbook(excel_sheet)
    sheet = wb["Sheet1"]

    # Set up the SMTP server
    server = smtplib.SMTP("smtp.gmail.com", 587)
    server.ehlo()
    server.starttls()
    server.ehlo()

    # Login to the email account
    server.login(username, app_password)

    # Iterate through the rows in the sheet
    for row in range(2, sheet.max_row + 1):
        name = sheet.cell(row, 1).value
        email = sheet.cell(row, 2).value

        # Compose the email message
        message = MIMEText(f"Hi {name},\n\nThis is a reminder for the experiment TOMORROW at time am/pm, which will take place in Psych East Basement. Once you arrive at the basement, you can dial 6190 on the phone at the entrance of the Brain Imaging Center or email me.\n\nBest regards,\nAyra Bandeli")
        message["Subject"] = "REMINDER Letter Detection Experiment"
        message["From"] = "Ayra Bandeli <your_email_address@gmail.com>"
        message["To"] = email

        # Send the email
        server.send_message(message)

    # Close the SMTP server connection
    server.quit()

